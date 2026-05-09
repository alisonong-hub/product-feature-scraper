"""
scraper.py - Core scraping logic for the Product Feature Scraper.
Imported by streamlit_app.py. No UI or file I/O here.
"""

import requests
import re
import time
import io
from difflib import SequenceMatcher
from urllib.parse import urlparse
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JINA_BASE = "https://r.jina.ai/"

# Smart-quote and mojibake cleanup -- use ordinal keys, chr() values to stay ASCII-safe
_CHAR_FIX = str.maketrans({
    0x2018: chr(39),    # left single quote -> '
    0x2019: chr(39),    # right single quote -> '
    0x201c: chr(34),    # left double quote -> "
    0x201d: chr(34),    # right double quote -> "
    0x2013: chr(45),    # en dash -> -
    0x2014: chr(45),    # em dash -> -
    0xfffd: None,       # replacement char -> remove
})

# Mojibake sequences produced when UTF-8 bytes are misread as latin-1
# e.g. the en dash U+2013 (bytes E2 80 93) becomes the 3-char string "\xe2\x80\x93"
# which, when read as latin-1, appears as "â\x80\x93"
_MOJIBAKE = [
    ("\xe2\x80\x93", chr(45)),    # en dash
    ("\xe2\x80\x94", chr(45)),    # em dash
    ("\xe2\x80\x99", chr(39)),    # right single quote
    ("\xe2\x80\x9c", chr(34)),    # left double quote
    ("\xe2\x80\x9d", chr(34)),    # right double quote
    ("\xe2\x80\xa6", "..."),      # ellipsis
    ("\xe2\x80\x98", chr(39)),    # left single quote
]

def _clean_text(text: str) -> str:
    """Normalise smart quotes and fix mojibake artefacts."""
    if not text:
        return text
    # Pass 1: fix properly UTF-8 encoded text that was decoded as latin-1
    try:
        text = text.encode("latin-1").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    # Pass 2: fix any remaining mojibake byte sequences
    for bad, good in _MOJIBAKE:
        text = text.replace(bad, good)
    # Pass 3: translate known Unicode punctuation to ASCII equivalents
    return text.translate(_CHAR_FIX).strip()


# ─────────────────────────────────────────────────────────────────────────────
# SCRAPER CLASS
# ─────────────────────────────────────────────────────────────────────────────

class ProductScraper:
    _STOP_WORDS = {
        "k18","hair","the","and","for","with","from","into","your","that","this",
        "size","pack","set","kit","new","best","top","oz","ml","fl","ounce",
        "care","style","use","day","free","safe","non","after","while","each",
        "all","add","per","also",
    }
    _DEPRIORITISE_WORDS = {
        "salon","offer","deal","valued","trio","bundle","intro","pro service"
    }

    def __init__(self, config):
        self.config    = config
        # Normalise to root domain - strip any path the user may have pasted
        # e.g. https://amika.com/collections/all → https://amika.com
        raw_url = config["store_url"].strip().rstrip("/")
        parsed  = urlparse(raw_url)
        self.store_url = f"{parsed.scheme}://{parsed.netloc}" if parsed.netloc else raw_url
        self.session   = requests.Session()
        self.session.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-AU,en;q=0.9",
        })
        self._product_cache = None
        self._is_shopify    = None

    # ── Platform detection ────────────────────────────────────────────────────

    @property
    def is_shopify(self):
        if self._is_shopify is None:
            self._is_shopify = self._detect_shopify()
        return self._is_shopify

    def platform_label(self):
        if self.is_shopify:
            if self._load_shopify_catalogue():
                return "Shopify"
            return "Shopify (search mode - products.json unavailable)"
        return "Non-Shopify (Jina mode)"

    def _detect_shopify(self):
        # Check 1: standard products.json endpoint
        try:
            resp = self.session.get(
                f"{self.store_url}/products.json?limit=1", timeout=8
            )
            if resp.status_code == 200 and "products" in resp.json():
                return True
        except Exception:
            pass

        # Check 2: HTML fingerprint - some Shopify stores block products.json
        # but still load cdn.shopify.com assets in their page source
        try:
            resp = self.session.get(self.store_url, timeout=10)
            html = resp.text.lower()
            shopify_signals = [
                "cdn.shopify.com",
                "shopify.com/s/",
                "myshopify.com",
                "shopify.theme",
                '"shopify"',
            ]
            if any(sig in html for sig in shopify_signals):
                return True
        except Exception:
            pass

        return False

    # ── Jina AI Reader ────────────────────────────────────────────────────────

    def _fetch_via_jina(self, url, retries=1):
        last_err = None
        for attempt in range(retries + 1):
            try:
                resp = self.session.get(
                    JINA_BASE + url,
                    timeout=25,
                    headers={"Accept": "text/markdown", "X-No-Cache": "true"},
                )
                resp.raise_for_status()
                return resp.text
            except Exception as e:
                last_err = e
                if attempt < retries:
                    time.sleep(2)
        raise last_err

    def _parse_sections_from_markdown(self, markdown):
        """
        Parse markdown into sections keyed by header text.
        Recognises both # Markdown headers AND short plain-text labels
        (e.g. "how to use?" or "ingredients" that appear on some Shopify sites
        as styled text rather than proper headings).
        """
        sections, current_key, lines = {}, None, []

        def _is_plain_label(line):
            """True if line looks like a section label (short, no special leading chars)."""
            stripped = line.strip()
            if not stripped or len(stripped) > 60:
                return False
            if stripped.startswith(("#", "*", "-", ">", "!", "[", "|", "http")):
                return False
            # Must contain mostly letters (not a paragraph sentence)
            alpha = sum(c.isalpha() for c in stripped)
            return alpha / max(len(stripped), 1) > 0.5

        for line in markdown.split("\n"):
            h = re.match(r"^#{1,6}\s+(.+)$", line)
            if h:
                if current_key:
                    sections[current_key] = "\n".join(lines).strip()
                current_key = re.sub(r"\*+", "", h.group(1)).strip().lower().rstrip(":").strip()
                lines = []
            elif _is_plain_label(line) and not lines:
                # Plain-text label at the start of a new section (nothing collected yet
                # under current key means the previous header was empty - commit it)
                if current_key:
                    sections[current_key] = "\n".join(lines).strip()
                current_key = line.strip().lower().rstrip(":").strip()
                lines = []
            elif current_key is not None:
                lines.append(line)

        if current_key and lines:
            sections[current_key] = "\n".join(lines).strip()

        return {k: _clean_text(v) for k, v in sections.items()}

    # ── Scoring helpers ───────────────────────────────────────────────────────

    def _fuzzy_score(self, a, b):
        return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()

    def _keyword_score(self, search_term, product_title):
        search_text = re.sub(r"[^a-z0-9\s]", " ", search_term.lower())
        title_words = re.sub(r"[^a-z0-9\s]", " ", product_title.lower()).split()
        meaningful  = [w for w in title_words if len(w) >= 3 and w not in self._STOP_WORDS]
        if not meaningful:
            return 0.0
        return sum(1 for w in meaningful if w in search_text) / len(meaningful)

    def _retail_penalty(self, title):
        t = title.lower()
        return 0.15 if any(w in t for w in self._DEPRIORITISE_WORDS) else 0.0

    # ── Shopify product discovery ─────────────────────────────────────────────

    def _load_shopify_catalogue(self):
        if self._product_cache is not None:
            return self._product_cache
        products, page = [], 1
        while True:
            try:
                resp  = self.session.get(
                    f"{self.store_url}/products.json?limit=250&page={page}", timeout=15
                )
                batch = resp.json().get("products", [])
            except Exception:
                break
            if not batch:
                break
            products.extend(batch)
            page += 1
            time.sleep(0.3)
        self._product_cache = products
        return products

    def catalogue_size(self):
        """Return number of products loaded (triggers load if needed)."""
        return len(self._load_shopify_catalogue())

    def _find_shopify_product(self, search_term, upc=None):
        products = self._load_shopify_catalogue()

        if upc:
            upc_str = str(upc).strip().split(".")[0]
            for p in products:
                for v in p.get("variants", []):
                    if str(v.get("barcode", "")).strip() == upc_str:
                        return (
                            f"{self.store_url}/products/{p['handle']}",
                            p["title"], "HIGH", f"Matched by UPC {upc_str}",
                        )

        search_clean = search_term.strip().lower()
        for p in products:
            if p["title"].strip().lower() == search_clean:
                return (
                    f"{self.store_url}/products/{p['handle']}",
                    p["title"], "HIGH", "Exact title match",
                )

        best_score, best = 0, None
        for p in products:
            score = (
                self._keyword_score(search_term, p["title"])
                + 0.05 * self._fuzzy_score(search_term, p["title"])
                - self._retail_penalty(p["title"])
            )
            if score > best_score:
                best_score, best = score, p

        if best and best_score >= 0.65:
            display_score = min(best_score, 1.0)
            reason = f"Keyword match ({display_score:.0%})"
            if upc:
                reason += f" - UPC {upc} not found in AU store (may be a US barcode)"
            return (
                f"{self.store_url}/products/{best['handle']}",
                best["title"], "HIGH", reason,
            )

        if best and best_score >= 0.35:
            kw = self._keyword_score(search_term, best["title"])
            parts = []
            if upc:
                parts.append(f"UPC {upc} not in AU store")
            if kw < 0.5:
                parts.append(f"Low keyword overlap ({kw:.0%}) - may be a bundle or US-only listing")
            else:
                parts.append(f"Score {best_score:.0%} below confidence threshold")
            return (
                f"{self.store_url}/products/{best['handle']}",
                best["title"], "LOW", "; ".join(parts),
            )

        reason = "No matching product found"
        if upc:
            reason += f" (UPC {upc} not in catalogue)"
        return None, None, None, reason

    # ── Shopify sitemap catalogue (blocked products.json fallback) ────────────

    def _load_sitemap_catalogue(self):
        """Fetch sitemap-products.xml and return a deduplicated list of product slugs.
        Result is cached on self._sitemap_slugs after the first call."""
        if hasattr(self, "_sitemap_slugs"):
            return self._sitemap_slugs
        slugs = []
        try:
            from xml.etree import ElementTree as ET
            resp = self.session.get(
                f"{self.store_url}/sitemap-products.xml", timeout=10
            )
            if resp.status_code == 200:
                tree = ET.fromstring(resp.content)
                ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
                seen = set()
                for loc in tree.findall(".//sm:loc", ns):
                    url = loc.text or ""
                    m = re.search(r"/products/([^?#/]+)", url)
                    if m:
                        slug = m.group(1)
                        if slug not in seen:
                            seen.add(slug)
                            slugs.append(slug)
        except Exception:
            pass
        self._sitemap_slugs = slugs
        return slugs

    def _find_shopify_slug_product(self, search_term, upc=None, direct_url=None):
        """
        Match input title against all product slugs in the sitemap using
        fuzzy + keyword scoring. Falls back to direct slug construction if
        the sitemap is unavailable.
        """
        if direct_url:
            return direct_url, search_term, "HIGH", "Direct URL provided"

        brand = self.config.get("brand_name", "").strip().lower()

        def _to_slug(text):
            t = re.sub(r"\b[\d.]+\s*(ml|oz|g|l|fl|mm|cm)\b", " ", text, flags=re.IGNORECASE)
            t = re.sub(r"[^a-z0-9\s]", " ", t.lower())
            return re.sub(r"\s+", "-", t.strip()).strip("-")

        def _drop_brand(slug):
            if brand:
                return re.sub(rf"^{re.escape(brand)}-", "", slug).strip("-")
            return slug

        # ── Step 1: sitemap fuzzy match ────────────────────────────────────────
        sitemap_slugs = self._load_sitemap_catalogue()
        if sitemap_slugs:
            input_slug = _to_slug(search_term)
            input_nb   = _drop_brand(input_slug)

            # Keywords: meaningful words from the (no-brand, no-size) slug
            stop = self._STOP_WORDS
            input_words = set(
                w for w in re.split(r"[^a-z0-9]+", input_nb)
                if len(w) >= 3 and w not in stop
            )

            best_score, best_slug = 0.0, None
            for slug in sitemap_slugs:
                slug_nb = _drop_brand(slug)
                fuzzy   = self._fuzzy_score(input_nb, slug_nb)
                slug_words = set(re.split(r"[^a-z0-9]+", slug_nb))
                kw = len(input_words & slug_words) / max(len(input_words), 1)
                score = 0.5 * fuzzy + 0.5 * kw
                if score > best_score:
                    best_score, best_slug = score, slug

            if best_slug and best_score >= 0.45:
                url = f"{self.store_url}/products/{best_slug}"
                try:
                    r = self.session.head(url, timeout=5, allow_redirects=True)
                    if r.status_code == 200:
                        title = best_slug.replace("-", " ").title()
                        return url, title, "HIGH", f"Sitemap match ({min(best_score,1.0):.0%})"
                except Exception:
                    pass

        # ── Step 2: direct slug construction fallback ──────────────────────
        no_size = re.sub(r"\b[\d.]+\s*(ml|oz|g|l|fl|mm|cm)\b", " ", search_term, flags=re.IGNORECASE)
        no_size = re.sub(r"\s+", " ", no_size).strip()

        def _drop_brand_str(t):
            if brand:
                return re.sub(rf"^{re.escape(brand)}\s+", "", t, flags=re.IGNORECASE).strip()
            return t

        def _make_slug(text):
            t = re.sub(r"[^a-z0-9\s]", " ", text.lower())
            return re.sub(r"\s+", "-", t.strip()).strip("-")

        candidates = list(dict.fromkeys(
            f"{self.store_url}/products/{_make_slug(v)}"
            for v in [search_term, no_size, _drop_brand_str(search_term), _drop_brand_str(no_size)]
            if v.strip()
        ))
        for url in candidates:
            try:
                r = self.session.head(url, timeout=8, allow_redirects=True)
                if r.status_code == 200:
                    slug = url.split("/products/")[-1]
                    return url, slug.replace("-", " ").title(), "HIGH", "Slug construction (verified 200)"
            except Exception:
                continue

        tried = ", ".join(c.split("/products/")[-1] for c in candidates)
        return None, None, None, (
            f"No match found (sitemap fuzzy + slug construction tried: {tried}) -- "
            "add a direct URL in column C"
        )

        # ── Non-Shopify discovery ─────────────────────────────────────────────────

    def _find_jina_product(self, search_term, upc=None, direct_url=None):
        if direct_url:
            return direct_url, search_term, "HIGH", "Direct URL provided"

        pattern = self.config.get("search_url_pattern", "")
        if not pattern:
            return (
                None, None, None,
                "No direct URL or search_url_pattern set. "
                "Add a product URL in column C of your input file.",
            )

        # Strip volume/size suffixes before URL-encoding - many retailers
        # (e.g. Sephora AU) return empty results when size info is in the query.
        query_clean = re.sub(r"\b[\d.]+\s*(ml|oz|g|l|fl|mm|cm)\b", " ", search_term, flags=re.IGNORECASE)
        query_clean = re.sub(r"[^a-z0-9\s]", " ", query_clean.lower())
        query_clean = re.sub(r"\s+", " ", query_clean).strip()
        search_url = pattern.replace("{query}", requests.utils.quote(query_clean))
        try:
            markdown = self._fetch_via_jina(search_url)
        except Exception as e:
            return None, None, None, f"Search page fetch failed: {e}"

        # Extract all product URLs directly from the raw markdown text.
        # This handles both:
        #   - Standard links:  [Product Title](https://site.com/products/handle)
        #   - Nested image-links used by Sephora AU and similar sites:
        #       [![alt text](https://cdn.../image.jpg)](https://site.com/products/handle)
        # The old [text](url) regex only captured the CDN image URL (inner parens)
        # in the nested format, missing the actual product URL (outer parens entirely).
        all_urls = re.findall(r"https?://[^\s\)\]\"'<>]+", markdown)
        store_domain = urlparse(self.store_url).netloc  # e.g. www.sephora.com.au
        _IMAGE_EXT = re.compile(r"\.(jpg|jpeg|png|gif|webp|svg|ico|avif)(\?|#|$)", re.IGNORECASE)
        product_urls = list(dict.fromkeys(
            u.rstrip(".,;") for u in all_urls
            if "/products/" in u
            and urlparse(u).netloc == store_domain   # same domain, not CDN
            and not _IMAGE_EXT.search(u)             # not an image file
        ))

        if not product_urls:
            return None, None, None, "No product links found in search results - add a direct URL in column C"

        def _title_to_slug(text):
            """Convert a product title to a URL-slug for comparison."""
            t = text.lower()
            t = re.sub(r"\b\d+\s*(ml|oz|g|l|fl)\b", "", t)   # strip sizes (50ml, 6.7oz)
            t = re.sub(r"[^a-z0-9\s]", " ", t)
            t = re.sub(r"\s+", "-", t.strip()).strip("-")
            return t

        def _slug_from_url(url):
            """Extract the product handle from a /products/{handle} URL."""
            m = re.search(r"/products/([^/?#]+)", url)
            return m.group(1) if m else ""

        input_slug = _title_to_slug(search_term)

        # Score every /products/ URL by fuzzy-matching input slug against URL slug
        seen = {}
        for url in product_urls:
            url_slug = _slug_from_url(url)
            if not url_slug:
                continue
            # Primary: slug similarity (clean, no noise)
            slug_score = self._fuzzy_score(input_slug, url_slug)
            # Secondary: keyword overlap as tiebreaker
            kw_score   = self._keyword_score(search_term, url_slug.replace("-", " "))
            score      = 0.8 * slug_score + 0.2 * kw_score
            if url not in seen or score > seen[url]:
                seen[url] = score

        if not seen:
            return None, None, None, "No product links found in search results - add a direct URL in column C"

        best_url   = max(seen, key=seen.get)
        best_score = seen[best_url]
        best_slug  = _slug_from_url(best_url)
        # Derive a readable title from the URL slug (actual title scraped from product page later)
        matched_title = best_slug.replace("-", " ").title()

        if best_score >= 0.65:
            return best_url, matched_title, "HIGH", f"Slug match ({min(best_score, 1.0):.0%})"

        return None, None, None, f"No confident match (best slug: '{best_slug}' at {min(best_score, 1.0):.0%}) - add a direct URL in column C"

    # ── Unified product finder ────────────────────────────────────────────────

    def find_product(self, search_term, upc=None, direct_url=None):
        if self.is_shopify and self._load_shopify_catalogue():
            # Full Shopify catalogue available - use fast in-memory matching
            return self._find_shopify_product(search_term, upc)
        elif self.is_shopify:
            # Shopify detected but products.json is blocked.
            # Step 1: try slug construction - fast, no search page needed.
            result = self._find_shopify_slug_product(search_term, upc, direct_url)
            if result[0]:
                return result
            # Step 2: fall back to Jina search as last resort.
            auto_pattern = f"{self.store_url}/search?q={{query}}&type=product"
            config_with_search = {**self.config, "search_url_pattern": auto_pattern}
            original_config = self.config
            self.config = config_with_search
            result = self._find_jina_product(search_term, upc, direct_url)
            self.config = original_config
            return result
        else:
            return self._find_jina_product(search_term, upc, direct_url)

    # ── HTML content extraction ───────────────────────────────────────────────

    def _fetch_html(self, url):
        resp = self.session.get(url, timeout=10)
        resp.raise_for_status()
        # Return raw bytes so BeautifulSoup/lxml reads the <meta charset> tag
        # and decodes correctly. Avoids requests guessing latin-1 and producing mojibake.
        return resp.content

    def _parse_html_sections(self, html):
        # html may be bytes (from _fetch_html) or str -- handle both
        soup = BeautifulSoup(html, "lxml")
        sections = {}

        # Pass 1: look inside a configured accordion selector (e.g. K18)
        accordion = soup.select_one(self.config.get("accordion_selector", ".accordion"))
        if accordion:
            for details in accordion.find_all("details"):
                summary = details.find("summary")
                if not summary:
                    continue
                parts = []
                for node in summary.children:
                    if not getattr(node, "name", None):
                        t = str(node).strip()
                        if t:
                            parts.append(t)
                    elif node.name not in ("img", "span", "svg"):
                        t = node.get_text(strip=True)
                        if t:
                            parts.append(t)
                header = " ".join(parts).strip().rstrip(":").strip()
                content_div = details.find(class_="details-content") or details
                content     = self._extract_html_text(content_div)
                if header:
                    sections[header.lower()] = content

        # Pass 2: scan ALL <details>/<summary> on the page (loveamika.com and similar)
        # This catches sections the accordion selector missed, including collapsed ones.
        for details in soup.find_all("details"):
            summary = details.find("summary")
            if not summary:
                continue
            header = summary.get_text(strip=True).lower().rstrip(":").strip()
            if not header or len(header) > 80 or header in sections:
                continue
            # Content = everything in <details> except the <summary>
            content_parts = []
            for child in details.children:
                if child == summary:
                    continue
                if hasattr(child, "get_text"):
                    t = child.get_text(separator="\n", strip=True)
                    if t:
                        content_parts.append(t)
            content = _clean_text("\n".join(content_parts).strip())
            if content:
                sections[header] = content

        if not sections:
            for tag in soup.find_all(["h2", "h3", "h4", "h5", "strong"]):
                header = tag.get_text(strip=True)
                if not header or len(header) > 60:
                    continue
                parts = []
                for sib in tag.find_next_siblings():
                    if sib.name in ("h2", "h3", "h4", "h5"):
                        break
                    t = sib.get_text(separator="\n", strip=True)
                    if t:
                        parts.append(t)
                if parts:
                    sections[header.lower()] = "\n".join(parts)

        return sections

    def _extract_html_text(self, element):
        lines = []

        def process(node):
            if not getattr(node, "name", None):
                return
            if node.name == "ol":
                for i, li in enumerate(node.find_all("li", recursive=False), 1):
                    lines.append(f"{i}. {li.get_text(separator=' ', strip=True)}")
            elif node.name == "ul":
                for li in node.find_all("li", recursive=False):
                    lines.append("• " + li.get_text(separator=" ", strip=True))
            elif node.name in ("p", "div", "span"):
                if node.find(["ol", "ul"]):
                    for child in node.children:
                        process(child)
                else:
                    t = node.get_text(separator=" ", strip=True)
                    if t:
                        lines.append(t)
            elif node.name in ("h2", "h3", "h4", "h5", "h6", "strong", "b"):
                t = node.get_text(strip=True)
                if t:
                    lines.append(t + ":")
            elif hasattr(node, "children"):
                for child in node.children:
                    process(child)

        root = element.find(class_="metafield-rich_text_field") or element
        for child in root.children:
            process(child)

        result = "\n".join(lines).strip()
        raw = result if result else element.get_text(separator="\n", strip=True)
        return _clean_text(raw)

    def _get_sections(self, url):
        """
        Return (sections_dict, method_label).
        For Shopify sites: parses HTML first. If all requested sections are
        found in the HTML, return immediately (fast path - no Jina needed).
        Only calls Jina when HTML sections are missing or empty, then merges
        results. This keeps K18-style sites fast while still filling collapsed
        accordion sections on sites like loveamika.com.
        For non-Shopify: Jina only.
        """
        requested = [s.lower().strip() for s in self.config.get("sections", [])]

        if self.is_shopify:
            html_sections = {}
            try:
                html_sections = self._parse_html_sections(self._fetch_html(url))
            except Exception:
                pass

            # Fast path: if HTML found content for every requested section, skip Jina.
            if html_sections and requested:
                covered = all(
                    any(
                        req in key or key in req or self._fuzzy_score(req, key) >= 0.6
                        for key in html_sections
                        if html_sections[key]
                    )
                    for req in requested
                )
                if covered:
                    return html_sections, "HTML"

            # Slow path: some sections missing - call Jina to fill gaps.
            try:
                markdown = self._fetch_via_jina(url)
                jina_sections = self._parse_sections_from_markdown(markdown)
                # Merge: prefer HTML values where present, fill blanks from Jina
                merged = {**jina_sections, **{k: v for k, v in html_sections.items() if v}}
                if any(merged.values()):
                    method = "HTML+Jina" if any(html_sections.values()) else "Jina"
                    return merged, method
            except Exception:
                pass

            if any(html_sections.values()):
                return html_sections, "HTML"

        markdown = self._fetch_via_jina(url)
        return self._parse_sections_from_markdown(markdown), "Jina"

    def _match_section(self, target, parsed):
        t = target.lower().strip()
        if t in parsed:
            return parsed[t]
        best_score, best_key = 0, None
        for key in parsed:
            score = self._fuzzy_score(t, key)
            if t in key or key in t:
                score = max(score, 0.8)
            if score > best_score:
                best_score, best_key = score, key
        return parsed[best_key] if best_key and best_score >= 0.6 else ""

    # ── Main scrape ───────────────────────────────────────────────────────────

    def scrape_product(self, product_name, upc=None, direct_url=None):
        result = {
            "input_name":    product_name,
            "matched_title": "",
            "product_url":   "",
            "status":        "",
            "note":          "",
        }
        for s in self.config["sections"]:
            result[s] = ""

        url, matched_title, confidence, reason = self.find_product(
            product_name, upc=upc, direct_url=direct_url
        )

        if not url:
            result["status"] = "NOT FOUND"
            result["note"]   = reason
            return result

        result["matched_title"] = matched_title or product_name
        result["product_url"]   = url

        if confidence == "LOW":
            result["status"] = "LOW CONFIDENCE"
            result["note"]   = reason
            return result

        try:
            sections, method = self._get_sections(url)
        except Exception as e:
            result["status"] = "ERROR"
            result["note"]   = str(e)
            return result

        for s in self.config["sections"]:
            result[s] = self._match_section(s, sections)

        result["status"] = "OK"
        result["note"]   = f"{reason} · via {method}"
        return result


# ─────────────────────────────────────────────────────────────────────────────
# INPUT PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_input(file_obj, filename):
    """
    Parse an uploaded file (file-like object) into a list of (name, upc, url).
    Supports .xlsx, .xls, .csv.
    """
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        df = pd.read_excel(file_obj, header=None, dtype=str)
    elif ext == "csv":
        df = pd.read_csv(file_obj, header=None, dtype=str)
    else:
        raise ValueError(f"Unsupported file type: .{ext}")

    first = str(df.iloc[0, 0]).lower().strip()
    if any(kw in first for kw in ("product","name","title","sku","upc","barcode","item")):
        df = df.iloc[1:].reset_index(drop=True)

    UPC_RE = re.compile(r"\(UPC[:\s]+(\d+)\)", re.IGNORECASE)
    rows = []
    for _, row in df.iterrows():
        name = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        upc  = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else None
        url  = str(row.iloc[2]).strip() if len(row) > 2 and pd.notna(row.iloc[2]) else None

        if upc:
            upc = upc.split(".")[0]
            if upc.lower() in ("nan", "", "none"):
                upc = None
        if url and url.lower() in ("nan", "", "none"):
            url = None

        if name:
            m = UPC_RE.search(name)
            if m:
                if not upc:
                    upc = m.group(1)
                name = UPC_RE.sub("", name).strip(" ,|--")

        if name and name.lower() not in ("nan", "", "none"):
            rows.append((name, upc, url))

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER  (returns BytesIO - no disk writes)
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(results, sections):
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Features"

    hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    meta_fill = PatternFill("solid", fgColor="D9E1F2")
    ok_fill   = PatternFill("solid", fgColor="E2EFDA")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    err_fill  = PatternFill("solid", fgColor="FCE4D6")
    c_align   = Alignment(horizontal="center", vertical="top", wrap_text=True)
    l_align   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    border    = Border(*[Side(style="thin")] * 0,
                       left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"),  bottom=Side(style="thin"))

    columns = ["Input Name","Matched Title","Product URL","Status","Note"] + sections
    for ci, col in enumerate(columns, 1):
        c = ws.cell(1, ci, col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = c_align; c.border = border

    for ri, r in enumerate(results, 2):
        row = [
            r.get("input_name",""), r.get("matched_title",""),
            r.get("product_url",""), r.get("status",""), r.get("note",""),
        ] + [r.get(s,"") for s in sections]
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.alignment = l_align; c.border = border
            if ci == 4:
                if val == "OK":
                    c.fill = ok_fill;   c.font = Font(name="Calibri", bold=True, color="375623")
                elif val == "LOW CONFIDENCE":
                    c.fill = warn_fill; c.font = Font(name="Calibri", bold=True, color="7F6000")
                elif val not in ("", None):
                    c.fill = err_fill;  c.font = Font(name="Calibri", bold=True, color="9C0006")
            elif ci <= 5:
                c.fill = meta_fill

    widths = {1:35, 2:35, 3:55, 4:14, 5:55}
    for i in range(len(sections)):
        widths[6+i] = 55
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    for ri in range(2, len(results)+2):
        ws.row_dimensions[ri].height = 80
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
